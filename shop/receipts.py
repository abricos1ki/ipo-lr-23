"""
Генерация чека заказа в формате Excel (.xlsx) с использованием openpyxl.
Лабораторная работа №19 "Работа с почтовым сервером и сторонними библиотеками".
"""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


def build_order_receipt(order):
    """
    Строит Excel-чек для переданного заказа (объект модели Order)
    и возвращает его содержимое в виде байтов (BytesIO).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Чек"

    bold = Font(bold=True)
    title_font = Font(bold=True, size=14)
    center = Alignment(horizontal="center")

    # Заголовок чека
    ws.merge_cells("A1:D1")
    ws["A1"] = "Интернет-магазин товаров для йоги"
    ws["A1"].font = title_font
    ws["A1"].alignment = center

    ws["A2"] = "Чек по заказу №"
    ws["B2"] = order.pk
    ws["A3"] = "Дата оформления:"
    ws["B3"] = order.created_at.strftime("%d.%m.%Y %H:%M")
    ws["A4"] = "Покупатель:"
    ws["B4"] = order.user.username
    ws["A5"] = "Адрес доставки:"
    ws["B5"] = order.delivery_address

    for row in (2, 3, 4, 5):
        ws[f"A{row}"].font = bold

    # Заголовки таблицы товаров
    header_row = 7
    headers = ["Товар", "Цена, BYN", "Количество", "Сумма, BYN"]
    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_index, value=header)
        cell.font = bold

    # Строки с товарами
    row = header_row + 1
    for item in order.items.select_related("product").all():
        ws.cell(row=row, column=1, value=item.product.name)
        ws.cell(row=row, column=2, value=float(item.price))
        ws.cell(row=row, column=3, value=item.quantity)
        ws.cell(row=row, column=4, value=float(item.item_cost()))
        row += 1

    # Итоговая строка
    total_row = row + 1
    ws.cell(row=total_row, column=3, value="Итого:").font = bold
    ws.cell(row=total_row, column=4, value=float(order.total_cost())).font = bold

    # Автоматическая ширина колонок (примерная)
    column_widths = [35, 14, 14, 14]
    for col_index, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(col_index)].width = width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
